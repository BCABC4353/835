"""
Fair Health Rates Parser

Parses rate data from Fair Health Excel files or Google Sheets and provides lookup
functionality for rates based on ZIP code, HCPCS code, and optional service date.

Supports loading from:
- Local Excel files (.xlsx)
- Google Sheets (public or shared with "Anyone with the link")

Column mapping from source:
- "Enter the location where you will be receiving or have received medical care" = ZIP Code
- "Date (GMT)" = date
- "Enter a Procedure Code or Keyword" = HCPCS
- "Out of Network" = Rate 1
- "In-Network" = Rate 2
"""

import csv
import io
import logging
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple, Union

logger = logging.getLogger(__name__)

# Try to import requests for Google Sheets support
try:
    import requests
except ImportError:
    requests = None

# Pre-compiled regex patterns for normalization (performance optimization)
_RE_NON_ALNUM = re.compile(r"[^A-Z0-9]")
_RE_NON_DIGIT = re.compile(r"[^0-9]")
_RE_CURRENCY = re.compile(r"[$,]")

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Google Sheets URL patterns
_RE_GOOGLE_SHEET_URL = re.compile(r"https?://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)")
_RE_GOOGLE_SHEET_ID = re.compile(r"^[a-zA-Z0-9_-]{20,}$")


def read_gsheet_file(filepath: str) -> Optional[str]:
    """
    Read a .gsheet shortcut file and extract the Google Sheet URL.

    .gsheet files are created by Google Drive for Desktop and contain
    JSON with a 'url' field pointing to the actual Google Sheet.

    Handles various scenarios:
    - Standard JSON format with 'url' field
    - Cloud placeholder files from Google Drive (tries multiple methods)
    - Different text encodings

    Args:
        filepath: Path to the .gsheet file

    Returns:
        The Google Sheet URL or None if not found
    """
    import json as json_module
    import os
    import subprocess
    import sys

    # First check if file exists
    if not os.path.exists(filepath):
        logger.error("Google Sheet file not found: %s", filepath)
        return None

    content = None
    read_method = None

    # Method 1: Try standard binary read
    try:
        with open(filepath, "rb") as f:
            raw_bytes = f.read()
        read_method = "binary"

        # Try different encodings
        for encoding in ["utf-8", "utf-8-sig", "utf-16", "latin-1"]:
            try:
                content = raw_bytes.decode(encoding).strip()
                if content:
                    break
            except (UnicodeDecodeError, UnicodeError):
                continue

    except PermissionError:
        logger.warning("Permission denied reading .gsheet file directly: %s", filepath)
    except OSError as e:
        if e.errno == 22:
            logger.warning("Cannot read .gsheet file directly (errno 22): %s - trying alternative methods", filepath)
        else:
            logger.warning("OSError reading .gsheet file: %s - %s", filepath, e)

    # Method 2: On Windows, try using 'type' command via subprocess
    # This sometimes works with Google Drive cloud files
    if not content and sys.platform == "win32":
        try:
            result = subprocess.run(
                ["cmd", "/c", "type", filepath],
                capture_output=True,
                timeout=10,
                check=False,
            )
            if result.returncode == 0 and result.stdout:
                for encoding in ["utf-8", "utf-8-sig", "cp1252", "latin-1"]:
                    try:
                        content = result.stdout.decode(encoding).strip()
                        if content:
                            read_method = "subprocess-type"
                            break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
        except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as e:
            logger.warning("Subprocess 'type' command failed: %s", e)

    # Method 3: Try PowerShell Get-Content (sometimes works with cloud files)
    if not content and sys.platform == "win32":
        try:
            result = subprocess.run(
                ["powershell", "-Command", f"Get-Content -Path '{filepath}' -Raw"],
                capture_output=True,
                timeout=10,
                check=False,
            )
            if result.returncode == 0 and result.stdout:
                for encoding in ["utf-8", "utf-8-sig", "utf-16", "latin-1"]:
                    try:
                        content = result.stdout.decode(encoding).strip()
                        if content:
                            read_method = "powershell"
                            break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
        except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as e:
            logger.warning("PowerShell Get-Content failed: %s", e)

    # If we still have no content, provide detailed error
    if not content:
        logger.error(
            "Cannot read .gsheet file: %s\n"
            "This is likely a Google Drive cloud-only placeholder file.\n\n"
            "SOLUTIONS:\n"
            "  1. In Google Drive folder, right-click the file and select 'Make available offline'\n"
            "  2. Or open the Google Sheet in your browser, copy the URL from the address bar,\n"
            "     and paste it directly into the RATES field in Settings\n"
            "  3. Or export the Google Sheet to Excel (.xlsx) and use the local file",
            filepath,
        )
        return None

    logger.debug("Read .gsheet file using method: %s", read_method)

    # Try to parse as JSON
    try:
        data = json_module.loads(content)
        if isinstance(data, dict):
            # Check for 'url' field (standard format)
            if "url" in data:
                logger.info("Found Google Sheet URL in .gsheet file")
                return data["url"]
            # Check for 'doc_id' field (alternate format)
            if "doc_id" in data:
                logger.info("Found Google Sheet doc_id in .gsheet file")
                return f"https://docs.google.com/spreadsheets/d/{data['doc_id']}"
    except json_module.JSONDecodeError:
        pass

    # If not JSON, check if the content itself contains a URL
    if "docs.google.com/spreadsheets" in content:
        match = _RE_GOOGLE_SHEET_URL.search(content)
        if match:
            logger.info("Extracted Google Sheet URL from .gsheet file content")
            return f"https://docs.google.com/spreadsheets/d/{match.group(1)}"

    # Check for just a sheet ID in the content
    content_stripped = content.strip()
    if _RE_GOOGLE_SHEET_ID.match(content_stripped):
        logger.info("Found Google Sheet ID in .gsheet file")
        return f"https://docs.google.com/spreadsheets/d/{content_stripped}"

    logger.error(
        "Could not find Google Sheet URL in .gsheet file: %s\n"
        "File content (first 200 chars): %s\n\n"
        "The file was readable but doesn't contain a valid Google Sheet reference.\n"
        "Try copying the URL directly from your browser instead.",
        filepath,
        content[:200] if content else "(empty)",
    )
    return None


def is_google_sheet(path: str) -> bool:
    """
    Check if a path is a Google Sheet URL, ID, or .gsheet file.

    Args:
        path: File path or URL to check

    Returns:
        True if this looks like a Google Sheet reference
    """
    if not path:
        return False
    path = path.strip()
    # Check for .gsheet file (Google Drive for Desktop shortcut)
    if path.lower().endswith(".gsheet"):
        return True
    # Check for Google Sheets URL
    if "docs.google.com/spreadsheets" in path:
        return True
    # Check for standalone sheet ID (long alphanumeric string)
    if _RE_GOOGLE_SHEET_ID.match(path):
        return True
    return False


def extract_google_sheet_id(path: str) -> Optional[str]:
    """
    Extract the Google Sheet ID from a URL or return the ID if already provided.

    Args:
        path: Google Sheet URL or ID

    Returns:
        The sheet ID or None if not a valid Google Sheet reference
    """
    if not path:
        return None
    path = path.strip()

    # Try to extract from URL
    match = _RE_GOOGLE_SHEET_URL.search(path)
    if match:
        return match.group(1)

    # Check if it's already a sheet ID
    if _RE_GOOGLE_SHEET_ID.match(path):
        return path

    return None


def extract_google_sheet_gid(path: str) -> str:
    """
    Extract the gid (tab ID) from a Google Sheet URL.

    Args:
        path: Google Sheet URL

    Returns:
        The gid string, or "0" if not found (defaults to first tab)
    """
    if not path:
        return "0"

    # Look for gid parameter in URL (e.g., ?gid=123456 or #gid=123456)
    gid_match = re.search(r"[?&#]gid=(\d+)", path)
    if gid_match:
        return gid_match.group(1)

    return "0"


def get_google_sheet_csv_url(sheet_id: str, gid: str = "0") -> str:
    """
    Get the CSV export URL for a Google Sheet.

    Args:
        sheet_id: The Google Sheet ID
        gid: The sheet/tab ID within the spreadsheet (default "0" for first sheet)

    Returns:
        URL that returns the sheet data as CSV
    """
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


# Headers to use when fetching Google Sheets (helps avoid getting HTML login pages)
_GOOGLE_SHEETS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/csv,text/plain,*/*",
}


def normalize_hcpcs(hcpcs: Any) -> Optional[str]:
    """
    Normalize HCPCS code: uppercase, strip whitespace, remove non-alphanumeric.
    Returns None if input is empty or invalid.
    """
    if hcpcs is None:
        return None
    s = str(hcpcs).strip().upper()
    if not s or s.lower() in ("undefined", "null", "none", "n/a", ""):
        return None
    # Remove non-alphanumeric characters
    s = _RE_NON_ALNUM.sub("", s)
    return s if s else None


def normalize_zip(zip_code: Any) -> Optional[int]:
    """
    Normalize ZIP code to 5-digit integer.
    Handles ZIP+4 format (12345-6789) and string inputs.
    Returns None if invalid.
    """
    if zip_code is None:
        return None
    s = str(zip_code).strip()
    if not s or s.lower() in ("undefined", "null", "none", "n/a", ""):
        return None
    # Handle ZIP+4 format
    if "-" in s:
        s = s.split("-")[0]
    # Extract first 5 digits
    digits = _RE_NON_DIGIT.sub("", s)
    if len(digits) >= 5:
        return int(digits[:5])
    elif digits:
        return int(digits)
    return None


def normalize_rate(rate: Any) -> Union[int, float, None]:
    """
    Normalize rate value.
    Returns numeric value or None (filters out N/A and non-numeric values).
    """
    if rate is None:
        return None
    s = str(rate).strip()
    if not s:
        return None
    # Remove currency symbols first, then check for undefined
    cleaned = _RE_CURRENCY.sub("", s).strip()
    # Filter out all non-numeric placeholder values
    if not cleaned or cleaned.lower() in ("undefined", "null", "none", "", "n/a", "error"):
        return None
    # Try to parse as number
    try:
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        # Non-numeric value - return None
        return None


def format_date(value: Any) -> str:
    """
    Format date as MM/DD/YY with leading zeros.
    """
    if value is None or value == "":
        return ""

    if isinstance(value, datetime):
        return value.strftime("%m/%d/%y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%y")

    s_value = str(value).strip()

    # Try common date formats
    date_formats = [
        ("%Y%m%d", r"^\d{8}$"),
        ("%Y-%m-%d", r"^\d{4}-\d{2}-\d{2}$"),
        ("%m/%d/%Y", r"^\d{1,2}/\d{1,2}/\d{4}$"),
        ("%m/%d/%y", r"^\d{1,2}/\d{1,2}/\d{2}$"),
    ]

    for fmt, pattern in date_formats:
        if re.match(pattern, s_value):
            try:
                dt_obj = datetime.strptime(s_value, fmt)
                return dt_obj.strftime("%m/%d/%y")
            except ValueError:
                continue

    # Try ISO format with time
    try:
        dt_obj = datetime.fromisoformat(s_value.replace("Z", "+00:00"))
        return dt_obj.strftime("%m/%d/%y")
    except ValueError:
        pass

    return s_value


class FairHealthRates:
    """
    Manages Fair Health rate data with date-aware lookups.

    Stores rates in two structures:
    - rate_ranges: Historical rates with date ranges for service date lookups
    - current_rates: Most recent rates for quick lookups
    """

    def __init__(self):
        self.rate_ranges: Dict[Tuple[int, str], List[dict]] = {}
        self.current_rates: Dict[Tuple[int, str], dict] = {}
        self.load_stats = {"rows_processed": 0, "rows_skipped": 0, "unique_zips": set(), "unique_hcpcs": set()}

    def load_from_excel(self, filepath: str) -> dict:
        """
        Load rate data from Excel file.

        Returns dict with load statistics.
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row, normalizing whitespace
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                h = str(cell.value)
                # Replace newlines, carriage returns, tabs with spaces
                normalized = re.sub(r"[\r\n\t]+", " ", h)
                # Collapse multiple spaces into one
                normalized = re.sub(r"\s+", " ", normalized).strip()
                headers.append(normalized)
            else:
                headers.append("")

        logger.debug("Excel headers found: %s", headers)

        # Map column names
        col_map = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if "location" in header_lower and "medical care" in header_lower:
                col_map["zip"] = idx
                logger.debug("Matched ZIP column at index %d: %s", idx, header)
            elif "date" in header_lower and "gmt" in header_lower:
                col_map["date"] = idx
                logger.debug("Matched date column at index %d: %s", idx, header)
            elif "procedure code" in header_lower or "keyword" in header_lower:
                col_map["hcpcs"] = idx
                logger.debug("Matched HCPCS column at index %d: %s", idx, header)
            elif "out of network" in header_lower:
                col_map["rate1"] = idx
                logger.debug("Matched rate1 (Out of Network) column at index %d: %s", idx, header)
            elif "in-network" in header_lower or "in network" in header_lower:
                col_map["rate2"] = idx
                logger.debug("Matched rate2 (In-Network) column at index %d: %s", idx, header)

        # Collect raw data grouped by (zip, hcpcs)
        raw_data: Dict[Tuple[int, str], List[dict]] = {}

        # Limit column iteration to only the columns we need (avoids reading empty columns)
        max_col = max(col_map.values()) + 1 if col_map else None

        for row in ws.iter_rows(min_row=2, max_col=max_col):
            self.load_stats["rows_processed"] += 1

            values = [cell.value for cell in row]

            # Get values
            zip_val = values[col_map.get("zip", 0)] if "zip" in col_map else None
            date_val = values[col_map.get("date", 1)] if "date" in col_map else None
            hcpcs_val = values[col_map.get("hcpcs", 2)] if "hcpcs" in col_map else None
            rate1_val = values[col_map.get("rate1", 3)] if "rate1" in col_map else None
            rate2_val = values[col_map.get("rate2", 4)] if "rate2" in col_map else None

            # Normalize values
            zip_code = normalize_zip(zip_val)
            hcpcs = normalize_hcpcs(hcpcs_val)
            rate1 = normalize_rate(rate1_val)
            rate2 = normalize_rate(rate2_val)

            # Skip invalid rows
            if zip_code is None or hcpcs is None:
                self.load_stats["rows_skipped"] += 1
                continue

            # Skip rows with no valid rates
            if rate1 is None and rate2 is None:
                self.load_stats["rows_skipped"] += 1
                continue

            # Parse date
            if isinstance(date_val, datetime):
                entry_date = date_val.date()
            elif isinstance(date_val, date):
                entry_date = date_val
            else:
                # Try to parse string date
                try:
                    if date_val:
                        s = str(date_val).strip()
                        # Try ISO format
                        entry_date = datetime.fromisoformat(s.replace("Z", "+00:00")).date()
                    else:
                        entry_date = date.today()
                except ValueError:
                    entry_date = date.today()

            key = (zip_code, hcpcs)
            if key not in raw_data:
                raw_data[key] = []

            raw_data[key].append({"date": entry_date, "rate1": rate1, "rate2": rate2})

            self.load_stats["unique_zips"].add(zip_code)
            self.load_stats["unique_hcpcs"].add(hcpcs)

        wb.close()

        # Build rate ranges from raw data
        self._build_rate_ranges(raw_data)

        return {
            "rows_processed": self.load_stats["rows_processed"],
            "rows_skipped": self.load_stats["rows_skipped"],
            "unique_zips": len(self.load_stats["unique_zips"]),
            "unique_hcpcs": len(self.load_stats["unique_hcpcs"]),
            "rate_keys": len(self.rate_ranges),
        }

    def _fetch_google_sheet_content(self, sheet_id: str, gid: str, original_url: str) -> tuple:
        """
        Fetch Google Sheet content with auto-discovery of correct tab.

        If the requested gid returns empty and it was the default (gid=0),
        this method will try to discover which tab actually contains data
        by checking the sheet's HTML for available tab IDs.

        Args:
            sheet_id: The Google Sheet ID
            gid: The tab ID to try first
            original_url: The original URL provided by user (for error messages)

        Returns:
            Tuple of (content, csv_url, actual_gid)

        Raises:
            ValueError: If no tab with data can be found
            Exception: If the sheet cannot be accessed
        """
        csv_url = get_google_sheet_csv_url(sheet_id, gid)
        logger.info("Fetching rates from Google Sheet: %s (gid=%s)", csv_url, gid)

        try:
            response = requests.get(csv_url, headers=_GOOGLE_SHEETS_HEADERS, timeout=30)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            raise Exception(
                f"Failed to fetch Google Sheet. Make sure the sheet is shared with 'Anyone with the link'. "
                f"URL attempted: {csv_url} | Error: {e}"
            ) from e

        content = response.text
        content_stripped = content.strip()

        # Check if we got HTML instead of CSV (indicates auth/access issue)
        if content_stripped.startswith("<!DOCTYPE") or content_stripped.startswith("<html"):
            raise ValueError(
                f"Google Sheet returned HTML instead of CSV data. "
                f"This usually means the sheet is not shared publicly. "
                f"IMPORTANT: You must share the sheet via the Share button (not just 'Publish to web'). "
                f"Set 'Anyone with the link' to 'Viewer'. "
                f"URL attempted: {csv_url}"
            )

        # If content is empty and we used the default gid, try to discover correct tab
        if not content_stripped and gid == "0":
            logger.info("Tab gid=0 is empty, attempting to discover correct tab...")
            discovered_gid = self._discover_sheet_tab(sheet_id)

            if discovered_gid and discovered_gid != "0":
                logger.info("Discovered tab with data: gid=%s", discovered_gid)
                # Recursively try with discovered gid
                return self._fetch_google_sheet_content(sheet_id, discovered_gid, original_url)
            else:
                # No tab discovered, provide helpful error
                raise ValueError(
                    f"Google Sheet returned empty content. This could mean:\n"
                    f"  1. The sheet tab (gid={gid}) is empty or doesn't exist\n"
                    f"  2. The sheet is not shared publicly (use Share button, set 'Anyone with link' to 'Viewer')\n"
                    f"  3. Your URL may be pointing to the wrong tab\n\n"
                    f"URL attempted: {csv_url}\n\n"
                    f"TIP: Open the Google Sheet in your browser, navigate to the tab with your data,\n"
                    f"     then copy the FULL URL from the address bar (it should include ?gid=XXXXXXX).\n"
                    f"     Paste that complete URL into Settings."
                )

        if not content_stripped:
            raise ValueError(
                f"Google Sheet returned empty content. This could mean:\n"
                f"  1. The sheet tab (gid={gid}) is empty or doesn't exist\n"
                f"  2. The sheet is not shared publicly (use Share button, set 'Anyone with link' to 'Viewer')\n"
                f"  3. Your URL may be pointing to the wrong tab\n\n"
                f"URL attempted: {csv_url}\n\n"
                f"TIP: Make sure your Google Sheet URL includes the correct gid parameter for the tab with data."
            )

        return content, csv_url, gid

    def _discover_sheet_tab(self, sheet_id: str) -> Optional[str]:
        """
        Attempt to discover which tab contains data by checking the sheet's HTML.

        Google Sheets HTML contains JavaScript with tab (gid) information.
        This method tries to extract available gids and find one with data.

        Args:
            sheet_id: The Google Sheet ID

        Returns:
            A gid string if a tab with data is found, None otherwise
        """
        try:
            # Fetch the sheet's HTML page to find available tabs
            html_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
            response = requests.get(html_url, headers=_GOOGLE_SHEETS_HEADERS, timeout=15)

            if response.status_code != 200:
                return None

            html_content = response.text

            # Look for gid patterns in the HTML/JavaScript
            # Google Sheets embeds tab info in various formats
            gid_pattern = re.compile(r'["\']?gid["\']?\s*[:=]\s*["\']?(\d+)["\']?', re.IGNORECASE)
            found_gids = set(gid_pattern.findall(html_content))

            # Remove "0" since we already tried it
            found_gids.discard("0")

            if not found_gids:
                logger.info("No additional tabs discovered in sheet HTML")
                return None

            logger.info("Found potential tabs: %s", list(found_gids)[:5])

            # Try each discovered gid until we find one with data
            for test_gid in sorted(found_gids, key=int):  # Try in numeric order
                try:
                    test_url = get_google_sheet_csv_url(sheet_id, test_gid)
                    test_response = requests.get(test_url, headers=_GOOGLE_SHEETS_HEADERS, timeout=10)
                    if test_response.status_code == 200:
                        test_content = test_response.text.strip()
                        if test_content and not test_content.startswith("<!DOCTYPE"):
                            # Found a tab with data!
                            logger.info("Tab gid=%s has content (%d chars)", test_gid, len(test_content))
                            return test_gid
                except requests.exceptions.RequestException:
                    continue

            return None

        except Exception as e:
            logger.debug("Tab discovery failed: %s", e)
            return None

    def load_from_google_sheet(self, sheet_url_or_id: str) -> dict:
        """
        Load rate data from a Google Sheet.

        The sheet must be publicly accessible or shared with "Anyone with the link".
        Uses CSV export which doesn't require API credentials.

        Args:
            sheet_url_or_id: Google Sheet URL or sheet ID

        Returns:
            dict with load statistics

        Raises:
            ImportError: If requests library is not installed
            ValueError: If the sheet URL/ID is invalid
            Exception: If the sheet cannot be accessed
        """
        if requests is None:
            raise ImportError("requests library is required for Google Sheets. Install with: pip install requests")

        sheet_id = extract_google_sheet_id(sheet_url_or_id)
        if not sheet_id:
            raise ValueError(f"Invalid Google Sheet URL or ID: {sheet_url_or_id}")

        # Extract gid (tab ID) from URL if present, otherwise use first tab
        gid = extract_google_sheet_gid(sheet_url_or_id)

        # Try to fetch the content, with auto-discovery of correct tab if gid=0 fails
        content, csv_url, gid = self._fetch_google_sheet_content(sheet_id, gid, sheet_url_or_id)

        # Build diagnostic info for error reporting (will be included in validation reports)
        diag_lines = []
        diag_lines.append(f"Response size: {len(content)} characters")

        # Capture first few lines of raw content for debugging
        content_lines = content.split("\n")
        diag_lines.append("First 3 lines of raw CSV data:")
        for i, line in enumerate(content_lines[:3]):
            # Truncate very long lines for readability
            display_line = line[:150] + "..." if len(line) > 150 else line
            diag_lines.append(f"  Line {i + 1}: {repr(display_line)}")

        reader = csv.reader(io.StringIO(content))

        # Get headers from first row
        raw_headers = next(reader, [])
        diag_lines.append(f"Raw headers ({len(raw_headers)} columns): {raw_headers[:10]}")
        if len(raw_headers) > 10:
            diag_lines.append(f"  ... and {len(raw_headers) - 10} more columns")

        # Normalize headers: strip whitespace, replace newlines/tabs with spaces, collapse multiple spaces
        headers = []
        for h in raw_headers:
            if h:
                # Replace newlines, carriage returns, tabs with spaces
                normalized = re.sub(r"[\r\n\t]+", " ", h)
                # Collapse multiple spaces into one
                normalized = re.sub(r"\s+", " ", normalized).strip()
                headers.append(normalized)
            else:
                headers.append("")

        diag_lines.append(f"Normalized headers: {headers[:10]}")

        # Map column names (same logic as Excel loader)
        col_map = {}
        matched_columns = []
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if "location" in header_lower and "medical care" in header_lower:
                col_map["zip"] = idx
                matched_columns.append(f"ZIP (col {idx}): '{header}'")
            elif "date" in header_lower and "gmt" in header_lower:
                col_map["date"] = idx
                matched_columns.append(f"Date (col {idx}): '{header}'")
            elif "procedure code" in header_lower or "keyword" in header_lower:
                col_map["hcpcs"] = idx
                matched_columns.append(f"HCPCS (col {idx}): '{header}'")
            elif "out of network" in header_lower:
                col_map["rate1"] = idx
                matched_columns.append(f"Out of Network (col {idx}): '{header}'")
            elif "in-network" in header_lower or "in network" in header_lower:
                col_map["rate2"] = idx
                matched_columns.append(f"In-Network (col {idx}): '{header}'")

        if matched_columns:
            diag_lines.append("Matched columns:")
            for mc in matched_columns:
                diag_lines.append(f"  - {mc}")
        else:
            diag_lines.append("No columns matched!")

        diag_lines.append(f"Column mapping result: {col_map}")

        # Log diagnostic info to console as well
        for line in diag_lines:
            logger.info(line)

        if not col_map:
            # Build comprehensive error message with all diagnostic info
            diag_lines.append("")
            diag_lines.append("EXPECTED COLUMN PATTERNS:")
            diag_lines.append("  - ZIP: header containing 'location' AND 'medical care'")
            diag_lines.append("  - Date: header containing 'date' AND 'gmt'")
            diag_lines.append("  - HCPCS: header containing 'procedure code' OR 'keyword'")
            diag_lines.append("  - Out of Network: header containing 'out of network'")
            diag_lines.append("  - In-Network: header containing 'in-network' or 'in network'")

            diagnostic_detail = "\n".join(diag_lines)
            raise ValueError(
                f"Could not find expected columns in Google Sheet.\n\n" f"DIAGNOSTIC INFORMATION:\n{diagnostic_detail}"
            )

        # Collect raw data grouped by (zip, hcpcs)
        raw_data: Dict[Tuple[int, str], List[dict]] = {}

        for row in reader:
            self.load_stats["rows_processed"] += 1

            # Get values (with safe indexing)
            def safe_get(idx_key):
                idx = col_map.get(idx_key)
                if idx is not None and idx < len(row):
                    return row[idx]
                return None

            zip_val = safe_get("zip")
            date_val = safe_get("date")
            hcpcs_val = safe_get("hcpcs")
            rate1_val = safe_get("rate1")
            rate2_val = safe_get("rate2")

            # Normalize values
            zip_code = normalize_zip(zip_val)
            hcpcs = normalize_hcpcs(hcpcs_val)
            rate1 = normalize_rate(rate1_val)
            rate2 = normalize_rate(rate2_val)

            # Skip invalid rows
            if zip_code is None or hcpcs is None:
                self.load_stats["rows_skipped"] += 1
                continue

            # Skip rows with no valid rates
            if rate1 is None and rate2 is None:
                self.load_stats["rows_skipped"] += 1
                continue

            # Parse date
            entry_date = date.today()
            if date_val:
                date_val = str(date_val).strip()
                # Try ISO format first
                try:
                    entry_date = datetime.fromisoformat(date_val.replace("Z", "+00:00")).date()
                except ValueError:
                    # Try other common formats
                    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y%m%d"]:
                        try:
                            entry_date = datetime.strptime(date_val, fmt).date()
                            break
                        except ValueError:
                            continue

            key = (zip_code, hcpcs)
            if key not in raw_data:
                raw_data[key] = []

            raw_data[key].append({"date": entry_date, "rate1": rate1, "rate2": rate2})

            self.load_stats["unique_zips"].add(zip_code)
            self.load_stats["unique_hcpcs"].add(hcpcs)

        # Build rate ranges from raw data
        self._build_rate_ranges(raw_data)

        return {
            "rows_processed": self.load_stats["rows_processed"],
            "rows_skipped": self.load_stats["rows_skipped"],
            "unique_zips": len(self.load_stats["unique_zips"]),
            "unique_hcpcs": len(self.load_stats["unique_hcpcs"]),
            "rate_keys": len(self.rate_ranges),
            "source": "google_sheet",
        }

    def load(self, source: str) -> dict:
        """
        Load rate data from either a local Excel file, Google Sheet URL, or .gsheet file.

        Automatically detects the source type based on the input.

        Args:
            source: Either a local file path (.xlsx/.gsheet) or Google Sheet URL/ID

        Returns:
            dict with load statistics

        Raises:
            ValueError: If source type cannot be determined
            Exception: If loading fails
        """
        if not source:
            raise ValueError("No rate source specified")

        source = source.strip()

        # Check if it's a .gsheet file (Google Drive for Desktop shortcut)
        if source.lower().endswith(".gsheet"):
            logger.info("Detected .gsheet file, reading Google Sheet URL from: %s", source)
            sheet_url = read_gsheet_file(source)
            if not sheet_url:
                raise ValueError(f"Could not read Google Sheet URL from .gsheet file: {source}")
            logger.info("Found Google Sheet URL: %s", sheet_url)
            return self.load_from_google_sheet(sheet_url)

        # Check if it's a Google Sheet URL or ID
        if is_google_sheet(source):
            logger.info("Detected Google Sheet source")
            return self.load_from_google_sheet(source)

        # Otherwise treat as local file
        if not source.lower().endswith(".xlsx"):
            logger.warning("Source doesn't end with .xlsx, attempting to load as Excel anyway: %s", source)

        logger.info("Loading from local Excel file: %s", source)
        return self.load_from_excel(source)

    def _build_rate_ranges(self, raw_data: Dict[Tuple[int, str], List[dict]]):
        """
        Build consolidated date ranges from daily entries.

        Consecutive days with the same rates are consolidated into ranges.
        """
        for key, entries in raw_data.items():
            # Sort by date
            entries.sort(key=lambda x: x["date"])

            ranges = []
            current_range = None

            for entry in entries:
                if current_range is None:
                    # Start new range
                    current_range = {
                        "start_date": entry["date"],
                        "end_date": entry["date"],
                        "rate1": entry["rate1"],
                        "rate2": entry["rate2"],
                    }
                elif (
                    entry["rate1"] == current_range["rate1"]
                    and entry["rate2"] == current_range["rate2"]
                    and entry["date"] <= current_range["end_date"] + timedelta(days=1)
                ):
                    # Extend current range
                    current_range["end_date"] = entry["date"]
                else:
                    # Save current range and start new one
                    ranges.append(current_range)
                    current_range = {
                        "start_date": entry["date"],
                        "end_date": entry["date"],
                        "rate1": entry["rate1"],
                        "rate2": entry["rate2"],
                    }

            # Don't forget the last range
            if current_range:
                ranges.append(current_range)

            self.rate_ranges[key] = ranges

            # Set current rate as the most recent
            if ranges:
                latest = ranges[-1]
                self.current_rates[key] = {
                    "rate1": latest["rate1"],
                    "rate2": latest["rate2"],
                    "as_of": latest["end_date"],
                }

    def get_zip_codes(self) -> List[int]:
        """
        Get all unique ZIP codes in the loaded rates data.

        Returns:
            Sorted list of ZIP codes as integers
        """
        return sorted(self.load_stats["unique_zips"])

    def get_rate(self, zip_code: int, hcpcs: str, service_date: Optional[date] = None) -> Optional[Tuple[Any, Any]]:
        """
        Look up rates for a given ZIP and HCPCS code.

        Args:
            zip_code: 5-digit ZIP code as integer
            hcpcs: HCPCS procedure code (will be normalized)
            service_date: Optional service date for historical lookup

        Returns:
            Tuple of (rate1, rate2) or None if not found
        """
        hcpcs_norm = normalize_hcpcs(hcpcs)
        if hcpcs_norm is None:
            return None

        key = (zip_code, hcpcs_norm)

        if service_date is None:
            # Return current rate
            if key in self.current_rates:
                cr = self.current_rates[key]
                return (cr["rate1"], cr["rate2"])
            return None

        # Look up by service date
        if key not in self.rate_ranges:
            return None

        for rate_range in self.rate_ranges[key]:
            if rate_range["start_date"] <= service_date <= rate_range["end_date"]:
                return (rate_range["rate1"], rate_range["rate2"])

        # If no exact match, return the most recent rate before the service date
        applicable = [r for r in self.rate_ranges[key] if r["end_date"] <= service_date]
        if applicable:
            latest = max(applicable, key=lambda x: x["end_date"])
            return (latest["rate1"], latest["rate2"])

        return None

    def get_rates_for_all_zips(self, hcpcs: str, service_date: Optional[date] = None) -> Dict[int, Tuple[Any, Any]]:
        """
        Look up rates for a given HCPCS code across all ZIP codes.

        Args:
            hcpcs: HCPCS procedure code (will be normalized)
            service_date: Optional service date for historical lookup

        Returns:
            Dict mapping ZIP code to (out_of_network, in_network) tuple
        """
        results = {}
        for zip_code in self.get_zip_codes():
            rate = self.get_rate(zip_code, hcpcs, service_date)
            if rate:
                results[zip_code] = rate
        return results

    def summary(self) -> str:
        """
        Generate a summary of loaded rate data.
        """
        lines = [
            "Fair Health Rates Summary",
            "=" * 40,
            f"Unique ZIP codes: {len(self.load_stats['unique_zips'])}",
            f"Unique HCPCS codes: {len(self.load_stats['unique_hcpcs'])}",
            f"Rate combinations: {len(self.rate_ranges)}",
            f"Rows processed: {self.load_stats['rows_processed']}",
            f"Rows skipped: {self.load_stats['rows_skipped']}",
            "",
            "Sample rates (first 10):",
            "-" * 40,
        ]

        for _i, (key, ranges) in enumerate(list(self.rate_ranges.items())[:10]):
            zip_code, hcpcs = key
            current = self.current_rates.get(key, {})
            rate1 = current.get("rate1", "")
            rate2 = current.get("rate2", "")
            as_of = current.get("as_of")
            as_of_str = format_date(as_of) if as_of else ""

            lines.append(f"  ZIP {zip_code} / {hcpcs}: ${rate1} (OON), ${rate2} (IN) as of {as_of_str}")

            if len(ranges) > 1:
                lines.append(f"    ({len(ranges)} date ranges)")
                for r in ranges[:3]:
                    start_str = format_date(r["start_date"])
                    end_str = format_date(r["end_date"])
                    lines.append(f"      {start_str} - {end_str}: ${r['rate1']} / ${r['rate2']}")
                if len(ranges) > 3:
                    lines.append(f"      ... and {len(ranges) - 3} more ranges")

        return "\n".join(lines)


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    if len(sys.argv) < 2:
        logger.error("Usage: python rates.py <excel_file>")
        sys.exit(1)

    filepath = sys.argv[1]
    rates = FairHealthRates()

    try:
        stats = rates.load_from_excel(filepath)
        logger.info("Loaded %d rate combinations", stats["rate_keys"])
        logger.info(rates.summary())
    except Exception as e:
        logger.error("Error loading rates: %s", e)
        import traceback

        traceback.print_exc()
        sys.exit(1)
